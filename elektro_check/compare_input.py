# -*- coding: utf-8 -*-
"""
Porovnani zadani (Input.xlsx, list FO) s daty vyextrahovanymi z DWG (export_full.xlsx).

List FO: sloupce A-O = puvodni kabely; CERVENA vypln = ruseny (zustava prazdne NEW),
jinak zustava a NOVE udaje jsou ve sloupcich S-Z. Kazdy kabel ma radky po vláknech (X1/X2).

Kontroly:
  1) Ruseni: kabel zruseny v zadani (vsechny vlakna cervena) <-> kabel REZERVA v DWG
  2) FROM/TO: nove Cab From/To v zadani <-> ODKUD/KAM v DWG (u kabelu co zustavaji)
  3) Inventura: kabely jen v zadani / jen v DWG

Pouziti: py compare_input.py [slozka]   (default final4)
"""
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

RED = "FFFF7C80"   # cervena vypln = ruseny

# sloupce listu FO (Excel 1-based)
C_CABFROM_OLD, C_FO, C_CABTO_OLD = 9, 11, 12
C_CABFROM_NEW, C_FO_NEW, C_CABTO_NEW = 19, 21, 22


def norm(s):
    if s is None:
        return ""
    s = " ".join(str(s).strip().upper().split())
    if " - " in s:                 # "OR5 - OR5" -> "OR5" (zadani vs DWG format)
        s = s.split(" - ")[0].strip()
    return s


def is_red(cell):
    f = cell.fill
    return f.patternType == "solid" and f.fgColor.rgb == RED


def read_spec(path):
    """fo_id -> {old_from, old_to, new_from, new_to, rows, red, kept}"""
    ws = load_workbook(path, data_only=True)["FO"]
    spec = {}
    for r in range(6, ws.max_row + 1):
        fo = ws.cell(r, C_FO).value
        fo = norm(fo)
        if not fo or "-" not in fo:
            continue
        e = spec.setdefault(fo, {"old_from": "", "old_to": "", "new_from": "",
                                 "new_to": "", "rows": 0, "red": 0, "kept": False})
        e["rows"] += 1
        red = is_red(ws.cell(r, C_FO))
        if red:
            e["red"] += 1
        if not e["old_from"]:
            e["old_from"] = norm(ws.cell(r, C_CABFROM_OLD).value)
            e["old_to"] = norm(ws.cell(r, C_CABTO_OLD).value)
        nf = norm(ws.cell(r, C_CABFROM_NEW).value)
        if not red and nf:
            e["kept"] = True
            e["new_from"] = nf
            e["new_to"] = norm(ws.cell(r, C_CABTO_NEW).value)
    for e in spec.values():
        e["cancelled"] = not e["kept"]      # vsechny vlakna ruseny / zadne NEW
    return spec


def read_dwg(path):
    """cable -> {odkud, kam, reserve}"""
    p = pd.read_excel(path, sheet_name="Propojeni", dtype=str).fillna("")
    dwg = {}
    for _, row in p.iterrows():
        c = norm(row.get("kabel"))
        if not c:
            continue
        e = dwg.setdefault(c, {"odkud": "", "kam": "", "reserve": False})
        if not e["odkud"]:
            e["odkud"] = norm(row.get("odkud_FROM"))
            e["kam"] = norm(row.get("kam_TO"))
        if norm(row.get("rezerva")) == "REZERVA":
            e["reserve"] = True
    return dwg


def norm_port(s):
    """'HUB3-PORT5 : X1' -> 'HUB3_PORT5' (sjednoceni se znacenim v DWG)."""
    if s is None:
        return ""
    s = str(s).split(":")[0]
    s = " ".join(s.strip().upper().split()).replace("-PORT", "_PORT")
    if " - " in s:
        s = s.split(" - ")[0].strip()
    return s


def read_spec_ports(path):
    """Mnozina HUB portu odkazovanych v zadani (Connect From/To, OLD i NEW)."""
    ws = load_workbook(path, data_only=True)["FO"]
    ports = set()
    for r in range(6, ws.max_row + 1):
        for c in (10, 13, 20, 23):     # J,M (OLD), T,W (NEW) = Connect From/To
            p = norm_port(ws.cell(r, c).value)
            if "HUB" in p and "PORT" in p:
                ports.add(p)
    return ports


def read_fo_in_or(path):
    """Kabel -> OR z listu 'FO in OR'."""
    ws = load_workbook(path, data_only=True)["FO in OR"]
    out = {}
    for r in range(3, ws.max_row + 1):
        fo = ws.cell(r, 2).value
        orr = ws.cell(r, 3).value
        if not fo or not orr:
            continue
        cable = norm(str(fo).split("\n")[0].split("(")[0])
        if cable and "-" in cable:
            out[cable] = norm(orr)
    return out


def read_dwg_devices(path):
    """Mnozina oznaceni zarizeni v DWG (vc. HUB portu)."""
    z = pd.read_excel(path, sheet_name="Zarizeni", dtype=str).fillna("")
    col = "VYPSANE_OZNACENI"
    return set(norm(t) for t in z[col]) if col in z.columns else set()


def run(folder):
    folder = Path(folder)
    spec = read_spec(folder / "Input.xlsx")
    dwg = read_dwg(folder / "export_full.xlsx")

    both = sorted(set(spec) & set(dwg))
    only_spec = sorted(set(spec) - set(dwg))
    only_dwg = sorted(set(dwg) - set(spec))

    rows = []
    ruseni_ok = ruseni_bad = 0
    ft_new = ft_old = ft_bad = 0
    for c in both:
        s, d = spec[c], dwg[c]
        ruseni_match = (s["cancelled"] == d["reserve"])
        ruseni_ok += ruseni_match
        ruseni_bad += not ruseni_match
        # FROM/TO jen u kabelu co zustavaji: DWG porovnej s OLD i NEW
        ft = ""
        if not s["cancelled"]:
            have = {d["odkud"], d["kam"]} - {""}
            new_set = {s["new_from"], s["new_to"]} - {""}
            old_set = {s["old_from"], s["old_to"]} - {""}
            if have and have == new_set:
                ft = "NEW OK"; ft_new += 1
            elif have and have == old_set:
                ft = "OLD (DWG nezmeneno?)"; ft_old += 1
            else:
                ft = "NESEDI"; ft_bad += 1
        rows.append({
            "kabel": c,
            "zadani_ruseny": "ANO" if s["cancelled"] else "",
            "DWG_rezerva": "ANO" if d["reserve"] else "",
            "ruseni_shoda": "OK" if ruseni_match else "!!! NESEDI",
            "zad_OLD_From": s["old_from"], "zad_OLD_To": s["old_to"],
            "zad_NEW_From": s["new_from"], "zad_NEW_To": s["new_to"],
            "DWG_ODKUD": d["odkud"], "DWG_KAM": d["kam"],
            "fromto_shoda": ft,
        })

    df = pd.DataFrame(rows)

    # --- Rozsireni 1: porty (Connect From/To) v zadani vs zarizeni v DWG ---
    spec_ports = read_spec_ports(folder / "Input.xlsx")
    dwg_devices = read_dwg_devices(folder / "export_full.xlsx")
    ports_missing = sorted(p for p in spec_ports if p not in dwg_devices)

    # --- Rozsireni 2: kabel -> OR (FO in OR) vs DWG ODKUD/KAM ---
    def or_match(spec_or, odkud, kam):
        ends = {odkud, kam}
        if spec_or in ends:
            return True
        base = spec_or.split("-OR")[0]      # 'N76-OR1' -> 'N76' (distributor ve skrini)
        return bool(base) and base in ends

    foinor = read_fo_in_or(folder / "Input.xlsx")
    or_ok, or_rows = 0, []
    for cable, orr in sorted(foinor.items()):
        if cable in dwg:
            d = dwg[cable]
            if or_match(orr, d["odkud"], d["kam"]):
                or_ok += 1
            else:
                or_rows.append({"kabel": cable, "zadani_OR": orr,
                                "DWG_ODKUD": d["odkud"], "DWG_KAM": d["kam"]})

    out = folder / "porovnani.xlsx"
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Detail", index=False)
        pd.DataFrame({"kabel_jen_v_zadani": only_spec}).to_excel(w, sheet_name="Jen_v_zadani", index=False)
        pd.DataFrame({"kabel_jen_v_DWG": only_dwg}).to_excel(w, sheet_name="Jen_v_DWG", index=False)
        pd.DataFrame({"port_v_zadani_chybi_v_DWG": ports_missing}).to_excel(
            w, sheet_name="Porty_chybi", index=False)
        pd.DataFrame(or_rows if or_rows else [{"kabel": "", "zadani_OR": "", "DWG_ODKUD": "", "DWG_KAM": ""}]).to_excel(
            w, sheet_name="OR_nesedi", index=False)

    print("=== POROVNANI zadani (FO) vs DWG ===")
    print("kabelu v zadani: {} | v DWG: {} | v obou: {}".format(len(spec), len(dwg), len(both)))
    print("jen v zadani (chybi v DWG): {}  | jen v DWG (chybi v zadani): {}".format(
        len(only_spec), len(only_dwg)))
    print()
    print("RUSENI (zadani cervene <-> DWG rezerva):  shoda {} / nesedi {}".format(ruseni_ok, ruseni_bad))
    print("FROM/TO (u zustavajicich kabelu):  DWG=NEW {} | DWG=OLD {} | nesedi {}".format(
        ft_new, ft_old, ft_bad))
    print("PORTY (HUB porty v zadani):  {} odkazovanych, {} chybi v DWG zarizenich".format(
        len(spec_ports), len(ports_missing)))
    if ports_missing:
        print("   chybi: {}".format(", ".join(ports_missing[:15])))
    print("KABEL->OR (FO in OR vs DWG):  shoda {} / nesedi {}".format(or_ok, len(or_rows)))
    for r in or_rows[:10]:
        print("   {} zadani_OR={} ale DWG {}/{}".format(r["kabel"], r["zadani_OR"], r["DWG_ODKUD"], r["DWG_KAM"]))
    print()
    bad = df[df.ruseni_shoda.str.contains("NESEDI")]
    if len(bad):
        print("--- Nesouhlasi ruseni (prvnich 15) ---")
        print(bad[["kabel", "zadani_ruseny", "DWG_rezerva"]].head(15).to_string(index=False))
    print()
    print("[XLS] {}".format(out))


if __name__ == "__main__":
    folder = sys.argv[1] if len(sys.argv) > 1 else r"D:\CLAUDE_TEST_DATA\elektro_check\final4"
    run(folder)
